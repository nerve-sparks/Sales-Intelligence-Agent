"""Header-agnostic ingest for arbitrary spreadsheets.

zoominfo_mapper.py handles exports that carry ZoomInfo's exact column names.
This module handles everything else - conference attendee lists, CRM exports,
hand-maintained prospect sheets - by resolving whatever headers a file happens
to use onto the same canonical fields, then synthesising the identity columns
the schema demands.

Why it exists: uploading a 1,009-row event meeting log ("AI4 Event.xlsx")
produced ZERO companies and reported "Complete" in 60ms. excel_pipeline
requires BOTH `ZoomInfo Company ID` and `Company Name` on every row and
silently `continue`s past anything else, so all 1,009 rows were dropped with no
warning. The file had `Company Name`, `Email Address`, `Designation` and
`LinkedIn` - genuinely useful prospect data the pipeline could not touch.

Three problems had to be solved, not just the column names:

  1. IDENTITY. company.zi_company_id and decision_maker.zi_person_id are both
     BIGINT NOT NULL, each half of a unique key with organisation_id. A file
     with no ZoomInfo ids still needs stable integers, and they must be
     STABLE - re-uploading the same sheet has to update rows rather than
     duplicate them. synthetic_bigint() derives them by hashing the strongest
     natural key available (domain, else normalised name).

  2. DOMAIN. search_signal_ingest skips any company whose company_domain is
     NULL, so a file without a Website column would ingest and then never be
     researched. domain_from_email() recovers it from contact email addresses
     ("Alexander@tacnode.io" -> "tacnode.io"), skipping free providers.

  3. HEADER POSITION AND SHEETS. read_tables() scans the first rows for the
     one that actually looks like a header (files often open with a title or
     blank row) and reads EVERY sheet, not just the active one - the AI4 file
     had eleven, and only the first was being read.
"""

import csv
import hashlib
import io
import re

import openpyxl

# How many leading rows to consider as a possible header row (see read_tables).
HEADER_SCAN_DEPTH = 8
# A header row must resolve at least this many canonical fields to be believed.
MIN_RECOGNISED_COLUMNS = 2

# Personal-email domains: seeing one tells us nothing about the company, so it
# must never become company_domain. An attendee list is full of them.
FREE_EMAIL_DOMAINS = {
    "gmail.com", "googlemail.com", "yahoo.com", "yahoo.co.uk", "yahoo.co.in",
    "hotmail.com", "hotmail.co.uk", "outlook.com", "live.com", "msn.com",
    "aol.com", "icloud.com", "me.com", "mac.com", "protonmail.com", "proton.me",
    "gmx.com", "gmx.de", "mail.com", "yandex.ru", "qq.com", "163.com", "126.com",
    "rediffmail.com", "zoho.com", "fastmail.com", "hey.com", "duck.com",
}

# canonical field -> header aliases, matched after _norm_header() (lowercased,
# punctuation and whitespace stripped). Longest-first within each list so
# "companywebsite" wins over "website" when both could match.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "company_name": (
        "companyname", "company", "accountname", "account", "organizationname",
        "organisationname", "organization", "organisation", "businessname",
        "employer", "firm", "firmname", "companyaccount",
    ),
    "website": (
        "companywebsite", "companydomain", "companyurl", "website", "domain",
        "weburl", "webaddress", "url", "web", "site", "homepage",
    ),
    "first_name": ("firstname", "givenname", "forename", "fname"),
    "last_name": ("lastname", "surname", "familyname", "lname"),
    "full_name": ("fullname", "contactname", "attendeename", "personname", "name"),
    "job_title": (
        "jobtitle", "designation", "title", "position", "jobrole", "role",
        "jobposition", "functionaltitle",
    ),
    "department": ("department", "function", "businessfunction", "dept"),
    "email": (
        "emailaddress", "workemail", "businessemail", "companyemail", "email",
        "emailid", "mail", "contactemail",
    ),
    "phone": (
        "directphonenumber", "phonenumber", "mobilenumber", "contactnumber",
        "phnumber", "telephone", "mobile", "phone", "tel", "ph",
    ),
    "linkedin": (
        "linkedincontactprofileurl", "linkedincompanyprofileurl", "linkedinurl",
        "linkedinprofile", "linkedin",
    ),
    "employees": ("employeecount", "numberofemployees", "headcount", "companysize", "employees", "size"),
    "revenue": ("annualrevenue", "revenueinusd", "revenue", "turnover", "sales"),
    "city": ("companycity", "city", "town"),
    "state": ("companystate", "state", "province", "region"),
    "country": ("companycountry", "country", "nation"),
    "industry": ("primaryindustry", "industry", "sector", "vertical", "companyindustry"),
    "founded": ("foundedyear", "yearfounded", "founded", "yearestablished", "established"),
    "ownership": ("ownershiptype", "ownership", "companytype", "ownershipstatus"),
    "funding": (
        "totalfunding", "totalfundingamount", "fundingamount", "funding",
        "totalraised", "capitalraised",
    ),
}

_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def _norm_header(value) -> str:
    return _NON_ALNUM.sub("", str(value).lower()) if value is not None else ""


def resolve_columns(header: list) -> dict[str, str]:
    """Maps canonical field -> the actual header string in this file.

    First alias wins, and aliases are ordered most-specific-first, so a sheet
    carrying both "Company Website" and "Website" resolves to the former. A
    header the file doesn't have simply stays absent from the result."""
    normalised = {}
    for raw in header:
        key = _norm_header(raw)
        if key and key not in normalised:
            normalised[key] = raw

    resolved: dict[str, str] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                resolved[field] = normalised[alias]
                break
    return resolved


def read_tables(filename: str, content: bytes) -> list[tuple[str, list[dict]]]:
    """Returns [(sheet_name, rows)] for every sheet that looks like a table.

    Reads ALL sheets: the AI4 event file had eleven ("Meetings", "Data Master",
    "50-100", "1K-5K", ...) and only the active one was ever read, so most of
    the workbook was invisible. Also finds the real header row rather than
    assuming row 1 - exported sheets routinely start with a title or a blank
    line, which would otherwise be read as the header and make every column
    unrecognisable."""
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return [("csv", list(csv.DictReader(io.StringIO(text))))]

    workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    tables: list[tuple[str, list[dict]]] = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        raw_rows = [row for row in sheet.iter_rows(values_only=True)]
        if not raw_rows:
            continue

        # Pick the header row: the best-resolving of the first few rows.
        best_index, best_score = None, 0
        for index, row in enumerate(raw_rows[:HEADER_SCAN_DEPTH]):
            score = len(resolve_columns(list(row)))
            if score > best_score:
                best_index, best_score = index, score
        if best_index is None or best_score < MIN_RECOGNISED_COLUMNS:
            continue  # not a table we can read - skip rather than guess

        header = [h if h is not None else "" for h in raw_rows[best_index]]
        body = raw_rows[best_index + 1:]
        rows = [
            dict(zip(header, row))
            for row in body
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        if rows:
            tables.append((sheet_name, rows))
    return tables


def synthetic_bigint(*parts) -> int:
    """A stable positive BIGINT derived from the given parts.

    company.zi_company_id and decision_maker.zi_person_id are BIGINT NOT NULL
    and each forms a unique key with organisation_id, so non-ZoomInfo rows need
    an integer identity. Hash-derived rather than sequential so it is STABLE:
    re-uploading the same sheet recomputes the same id and updates the existing
    row instead of inserting a duplicate.

    Masked to 63 bits to stay inside signed BIGINT and positive. Real ZoomInfo
    ids are ~9 digits while these are ~18-19, so the two spaces do not collide
    in practice."""
    digest = hashlib.sha1("|".join("" if p is None else str(p) for p in parts).encode()).digest()
    return int.from_bytes(digest[:8], "big") & 0x3FFF_FFFF_FFFF_FFFF


def domain_from_email(email) -> str | None:
    """Company domain recovered from a work email, or None for a personal one.

    This is what makes a contact-only sheet researchable at all:
    search_signal_ingest skips every company with company_domain IS NULL, so
    without this an attendee list would ingest and then be silently passed over
    by the research stage."""
    if not email or "@" not in str(email):
        return None
    domain = str(email).strip().lower().rsplit("@", 1)[-1]
    domain = domain.strip(" .<>,;")
    if not domain or "." not in domain or domain in FREE_EMAIL_DOMAINS:
        return None
    return domain


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in ("none", "nan", "n/a", "na", "-", "null"):
        return None
    return text


def _get(row: dict, columns: dict[str, str], field: str) -> str | None:
    header = columns.get(field)
    return _clean(row.get(header)) if header else None


def _revenue_thousands(value: str | None) -> int | None:
    """Ordinary sheets give dollars (or ranges). ZoomInfo's ingest path expects
    thousands-of-USD integers. Ranges like '$10M-$50M' cannot be converted and
    stay None so Revenue Range carries the display string instead."""
    if not value:
        return None
    text = value.strip().lower().replace(",", "").replace("$", "")
    if "-" in text or " to " in text or any(c.isalpha() for c in text if c not in "kmbt."):
        # Likely a range or labelled amount - leave for Revenue Range only.
        # Plain "50000000" / "50m" still convert below.
        if re.search(r"[a-z]", text) and not re.fullmatch(r"[\d.]+[kmbt]?", text):
            return None
    match = re.fullmatch(r"([\d.]+)([kmbt])?", text)
    if not match:
        return None
    amount = float(match.group(1))
    mult = {None: 1, "k": 1_000, "m": 1_000_000, "b": 1_000_000_000, "t": 1_000_000_000_000}[match.group(2)]
    dollars = int(amount * mult)
    return dollars // 1000 if dollars >= 1000 else dollars


def split_full_name(full_name: str | None) -> tuple[str | None, str | None]:
    """Splits a single name column into first/last. Attendee lists often carry
    one "Name" column where ZoomInfo exports carry two."""
    if not full_name:
        return None, None
    parts = full_name.split()
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])


def company_identity(row: dict, columns: dict[str, str]) -> tuple[str, str | None] | None:
    """Returns (company_name, company_domain) or None when the row names no
    company at all. Domain falls back to the email's domain - see
    domain_from_email for why that matters."""
    name = _get(row, columns, "company_name")
    if not name:
        return None
    domain = normalize_domain(_get(row, columns, "website"))
    if not domain:
        domain = domain_from_email(_get(row, columns, "email"))
    return name, domain


def normalize_domain(website) -> str | None:
    """Same normalisation as zoominfo_mapper.normalize_domain, duplicated here
    so this module stands alone; kept byte-identical so a domain ingested by
    either path produces the same company_domain and therefore the same
    identity."""
    if not website:
        return None
    domain = str(website).strip().lower()
    domain = re.sub(r"^https?://", "", domain)
    domain = re.sub(r"^www\.", "", domain)
    domain = domain.split("/")[0].strip()
    return domain or None


def normalize_company_name(name: str) -> str:
    """Identity key for name-only rows. Strips the legal suffixes and
    punctuation that make "Acme, Inc." and "Acme Inc" look like two companies."""
    text = str(name).lower()
    text = _NON_ALNUM.sub(" ", text)
    words = [w for w in text.split() if w not in _LEGAL_SUFFIXES]
    return " ".join(words) or text.strip()


_LEGAL_SUFFIXES = {
    "inc", "incorporated", "llc", "llp", "lp", "ltd", "limited", "plc", "corp",
    "corporation", "co", "company", "gmbh", "ag", "sa", "sas", "bv", "nv", "pty",
    "pvt", "private", "holdings", "group", "the",
}


# ── ADAPTER ──────────────────────────────────────────────────────────────────
# Emitted with ZoomInfo's exact header names so excel_pipeline.upsert_rows and
# zoominfo_mapper.build_*_row consume these rows unchanged. Adapting at the
# boundary keeps the whole downstream pipeline - and the existing ZoomInfo
# path that already ingested thousands of companies - completely untouched.
#
# These four are read with bracket access (row["..."]) rather than .get(), so
# every emitted row MUST carry them or ingestion raises KeyError.
_REQUIRED_KEYS = ("ZoomInfo Company ID", "ZoomInfo Contact ID", "Founded Year", "Primary Industry")


def looks_like_zoominfo_export(rows: list[dict]) -> bool:
    """True when the file already carries ZoomInfo's own identity column, in
    which case the untouched zoominfo_mapper path handles it."""
    return bool(rows) and "ZoomInfo Company ID" in rows[0]


def to_canonical_rows(filename: str, content: bytes) -> tuple[list[dict], dict]:
    """Converts any spreadsheet into ZoomInfo-shaped rows.

    Returns (rows, report). The report is what makes a bad upload legible
    instead of silent - it records, per sheet, which canonical fields were
    recognised and how many rows were dropped, so "0 companies ingested" can
    always be explained. The original failure here reported "Complete" with no
    warning after discarding 1,009 rows."""
    report = {
        "sheets": [], "rows_read": 0, "rows_usable": 0,
        "rows_without_company": 0, "companies": 0, "contacts": 0,
        "companies_without_domain": 0,
    }
    out: list[dict] = []
    seen_companies: set[int] = set()

    for sheet_name, rows in read_tables(filename, content):
        columns = resolve_columns(list(rows[0].keys()))
        sheet_stat = {
            "sheet": sheet_name, "rows": len(rows),
            "recognised_fields": sorted(columns), "dropped_no_company": 0,
        }
        report["rows_read"] += len(rows)

        for row in rows:
            identity = company_identity(row, columns)
            if identity is None:
                sheet_stat["dropped_no_company"] += 1
                report["rows_without_company"] += 1
                continue
            company_name, domain = identity

            # Domain is the stronger natural key: it merges "Acme, Inc." and
            # "Acme Inc" that a name-derived id would split into two companies.
            zi_company_id = synthetic_bigint("company", domain or normalize_company_name(company_name))

            first = _get(row, columns, "first_name")
            last = _get(row, columns, "last_name")
            if not first and not last:
                first, last = split_full_name(_get(row, columns, "full_name"))
            email = _get(row, columns, "email")

            # A row with no identifiable person still carries the company, so
            # it is kept - upsert_rows skips contact creation when the contact
            # id is absent, giving us the company either way.
            has_contact = bool(email or first or last)
            zi_person_id = (
                synthetic_bigint("contact", email or f"{zi_company_id}|{first}|{last}")
                if has_contact else None
            )

            out.append({
                "ZoomInfo Company ID": zi_company_id,
                "ZoomInfo Contact ID": zi_person_id,
                "Company Name": company_name,
                "Website": domain,
                "Founded Year": _get(row, columns, "founded"),
                # build_company_row maps this to company_status/is_verified,
                # and treats anything but "Yes" as OUT_OF_BUSINESS. Absence of
                # a ZoomInfo certification column is not evidence a company has
                # closed - without this every company from an ordinary
                # spreadsheet was stored as out of business.
                "Certified Active Company": "Yes",
                "Primary Industry": _get(row, columns, "industry"),
                "Employees": _get(row, columns, "employees"),
                "Employee Range": _get(row, columns, "employees"),
                # Revenue: ZoomInfo's column is "in 000s USD". Ordinary sheets
                # usually carry a display range ("$10M-$50M") or a raw dollar
                # figure. Prefer the range for display; when the cell is a plain
                # number, convert dollars -> thousands so build_company_row's
                # *1000 lands on the right magnitude.
                "Revenue (in 000s USD)": _revenue_thousands(_get(row, columns, "revenue")),
                "Revenue Range (in USD)": _get(row, columns, "revenue"),
                "Ownership Type": _get(row, columns, "ownership"),
                "Total Funding Amount (in 000s USD)": _revenue_thousands(_get(row, columns, "funding")),
                "Company City": _get(row, columns, "city"),
                "Company State": _get(row, columns, "state"),
                "Company Country": _get(row, columns, "country"),
                "First Name": first,
                "Last Name": last,
                "Job Title": _get(row, columns, "job_title"),
                "Department": _get(row, columns, "department"),
                "Email Address": email,
                "Direct Phone Number": _get(row, columns, "phone"),
                "LinkedIn Contact Profile URL": _get(row, columns, "linkedin"),
                "_source_sheet": sheet_name,
            })
            report["rows_usable"] += 1
            if has_contact:
                report["contacts"] += 1
            if zi_company_id not in seen_companies:
                seen_companies.add(zi_company_id)
                report["companies"] += 1
                if not domain:
                    report["companies_without_domain"] += 1

        report["sheets"].append(sheet_stat)
    return out, report


def report_warnings(report: dict) -> list[str]:
    """Human-readable warnings for IcpImportBatch.processing_warnings.

    Deliberately reports the GOOD path too ("read N rows across M sheets"),
    because the failure that motivated this module was indistinguishable from
    success: no warning, status Complete, zero companies."""
    name = report.get("file") or "file"
    warnings: list[str] = []
    sheets = report.get("sheets") or []
    if not sheets:
        warnings.append(
            f"{name}: no readable table found - no sheet had at least "
            f"{MIN_RECOGNISED_COLUMNS} recognisable columns (a company-name column is required)."
        )
        return warnings

    warnings.append(
        f"{name}: read {report['rows_read']} row(s) across {len(sheets)} sheet(s) "
        f"-> {report['companies']} company(ies), {report['contacts']} contact(s)."
    )
    if report["rows_without_company"]:
        warnings.append(
            f"{name}: skipped {report['rows_without_company']} row(s) with no company name."
        )
    # The one that actually changes what the user gets: research skips any
    # company with a NULL domain, so these ingest but are never enriched.
    if report["companies_without_domain"]:
        warnings.append(
            f"{name}: {report['companies_without_domain']} of {report['companies']} company(ies) "
            "have no website or work-email domain, so they cannot be researched for buying "
            "evidence and will score on contact access only."
        )
    unreadable = [s["sheet"] for s in sheets if not s["recognised_fields"]]
    if unreadable:
        warnings.append(f"{name}: sheet(s) with no recognisable columns: {', '.join(unreadable)}.")
    return warnings
