import io
import logging
from datetime import datetime, timezone
from uuid import UUID

import openpyxl
from sqlalchemy import delete, func, select, text, update
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.db import async_session_maker
from app.models import (
    BuyingEvent,
    Company,
    CompanyImportBatch,
    DecisionMaker,
    IcpImportBatch,
    LeadScore,
)
from app.services import company_enrichment, evidence_scorer, search_signal_ingest
from app.services import zoominfo_mapper as mapper
from app.services.offering_profile_service import ensure_offering_profile

COMPANY_UPDATE_COLS = [c for c in mapper.COMPANY_COLUMNS if c not in ("zi_company_id", "company_id", "organisation_id")]
DM_UPDATE_COLS = [c for c in mapper.DECISION_MAKER_COLUMNS if c not in ("zi_person_id", "organisation_id", "company_id")]

logger = logging.getLogger(__name__)


# A multi-row INSERT spends one bind parameter per column per row, and the
# driver caps how many one statement may carry. Inserting every row at once
# therefore breaks past a certain upload size - a 2,573-company file needs
# 2573 x 33 = 84,909 parameters and the failure arrives as a multi-megabyte
# dump of every placeholder.
#
# 32767, NOT PostgreSQL's documented 65535. asyncpg reports
#
#     asyncpg.exceptions._base.InterfaceError:
#     the number of query arguments cannot exceed 32767
#
# because the wire protocol counts parameters in a SIGNED int16. Sizing to the
# documented 65535 looks correct, passes a "under the limit?" check, and still
# fails at runtime - which is exactly what happened here on the first attempt
# at this fix.
#
# The limit was unreachable while only ZoomInfo exports could be ingested;
# table_mapper made arbitrary spreadsheets ingestable and immediately exceeded
# it on both tables at once - 8,338 contacts x 14 = 116,732.
#
# Chunk sizes are DERIVED from the column count rather than hardcoded, so
# adding a column to either model cannot silently reintroduce this. The 90%
# margin leaves room for the ON CONFLICT DO UPDATE clause's own parameters.
PG_MAX_BIND_PARAMS = 32_767


def _max_rows_per_insert(column_count: int) -> int:
    return max(1, int(PG_MAX_BIND_PARAMS * 0.9) // column_count)


COMPANY_INSERT_CHUNK = _max_rows_per_insert(len(mapper.COMPANY_COLUMNS))
DM_INSERT_CHUNK = _max_rows_per_insert(len(mapper.DECISION_MAKER_COLUMNS))


async def _upsert_companies(session: AsyncSession, company_rows: list[dict]) -> None:
    for start in range(0, len(company_rows), COMPANY_INSERT_CHUNK):
        chunk = company_rows[start : start + COMPANY_INSERT_CHUNK]
        stmt = pg_insert(Company).values(chunk)
        update_cols = {c: getattr(stmt.excluded, c) for c in COMPANY_UPDATE_COLS}
        stmt = stmt.on_conflict_do_update(index_elements=["organisation_id", "zi_company_id"], set_=update_cols)
        await session.execute(stmt)


async def _upsert_decision_makers(session: AsyncSession, dm_rows: list[dict]) -> None:
    for start in range(0, len(dm_rows), DM_INSERT_CHUNK):
        chunk = dm_rows[start : start + DM_INSERT_CHUNK]
        stmt = pg_insert(DecisionMaker).values(chunk)
        update_cols = {c: getattr(stmt.excluded, c) for c in DM_UPDATE_COLS}
        stmt = stmt.on_conflict_do_update(index_elements=["organisation_id", "zi_person_id"], set_=update_cols)
        await session.execute(stmt)


async def upsert_rows(session: AsyncSession, organisation_id: UUID, raw_rows: list[dict]) -> dict[int, UUID]:
    """Parses prospect rows and upserts ONLY identity + firmographic + contact
    data (brief item 14). External buying evidence (news/scoops/intent) no
    longer comes from spreadsheet columns - it all originates through Tavily
    research into buying_event, so build_intent_row/build_scoop_row/
    build_news_row are deliberately not called here.

    Returns a {zi_company_id: company_id} map for every company referenced.
    """
    seen_companies: dict[int, dict] = {}
    seen_dms: dict[int, dict] = {}
    dm_rows_no_id: list[dict] = []
    zi_to_company_id: dict[int, UUID] = {}

    for row in raw_rows:
        zi_company_id = mapper.parse_int(row.get("ZoomInfo Company ID"))
        if not zi_company_id or not row.get("Company Name"):
            continue  # orphaned contact row with no linked company

        if zi_company_id not in seen_companies:
            company_row = mapper.build_company_row(row, organisation_id)
            seen_companies[zi_company_id] = company_row
            zi_to_company_id[zi_company_id] = company_row["company_id"]

        # Multiple uploaded files can legitimately contain the same contact -
        # a single bulk INSERT can't ON CONFLICT DO UPDATE the same
        # (organisation_id, zi_person_id) row twice, so duplicates within this
        # batch are collapsed here (last occurrence wins). Rows with no
        # zi_person_id never conflict (NULL != NULL in a unique constraint).
        dm_row = mapper.build_decision_maker_row(row, organisation_id)
        zi_person_id = dm_row["zi_person_id"]
        if zi_person_id is not None:
            seen_dms[zi_person_id] = dm_row
        else:
            dm_rows_no_id.append(dm_row)

    await _upsert_companies(session, list(seen_companies.values()))
    await _upsert_decision_makers(session, list(seen_dms.values()) + dm_rows_no_id)
    await session.commit()

    return zi_to_company_id


async def run_pipeline(
    session: AsyncSession, organisation_id: UUID, raw_rows: list[dict]
) -> dict[int, UUID]:
    """The fast, synchronous half of a prospect upload: parse + upsert
    company/contact identity data only (brief section 4). No ICP, no signals,
    no scoring - external buying evidence (Tavily research) and scoring are
    the slow parts and run afterward in the background task, so the upload
    endpoint returns immediately. Returns {zi_company_id: company_id}.
    """
    return await upsert_rows(session, organisation_id, raw_rows)


async def _company_ids_for_batch(session: AsyncSession, import_batch_id: UUID) -> list[UUID]:
    """This batch's companies via the permanent membership table (item 5) -
    never Company.import_batch_id, which a later re-upload overwrites."""
    return list(
        (
            await session.execute(
                select(CompanyImportBatch.company_id).where(
                    CompanyImportBatch.import_batch_id == import_batch_id
                )
            )
        ).scalars().all()
    )


async def _cancel_requested(session: AsyncSession, import_batch_id: UUID) -> bool:
    return (
        await session.execute(
            select(IcpImportBatch.cancel_requested_at).where(IcpImportBatch.import_batch_id == import_batch_id)
        )
    ).scalar() is not None


async def score_companies_in_background(
    organisation_id: UUID, workspace_id: UUID, import_batch_id: UUID
) -> None:
    """Runs as a FastAPI BackgroundTask after the upload response is sent (own
    DB session). Scoped to THIS upload's companies via the membership table.

    The entire pipeline is wrapped so the batch NEVER stays permanently
    'pending' (brief item 6): on success it flips to 'complete' (or
    'complete_with_warnings' if some companies' research failed), and on an
    unhandled exception it flips to 'failed' with the error recorded. Per-stage
    counts (researched / research-failures / llm-failures / scoring-failures)
    and processing timestamps are persisted for observability.

    Checks POST .../cancel's cancel_requested_at before each of the two major
    stages (research, scoring) - cooperative/best-effort: a stage already in
    flight runs to completion (its own per-company work is cheap relative to
    a 1000-company batch), but the NEXT stage is skipped once cancellation is
    seen, so a cancelled job stops making progress promptly without needing
    per-company interruption plumbed through the concurrent research/scoring
    gather calls.
    """
    started = datetime.now(timezone.utc)
    warnings: list[str] = []
    research_summary: dict = {}
    counts = {label: 0 for label in evidence_scorer.STATUS_TO_COUNT}
    status = "failed"
    error: str | None = None
    log_ctx = {"job_id": str(import_batch_id), "stage": "job"}
    logger.info("job started", extra=log_ctx)
    print(f"\n{'='*80}\n[UPLOAD] Background scoring job STARTED\n"
          f"[UPLOAD]   organisation_id = {organisation_id}\n"
          f"[UPLOAD]   workspace_id    = {workspace_id}\n"
          f"[UPLOAD]   import_batch_id = {import_batch_id}\n"
          f"[UPLOAD]   started_at      = {started.isoformat()}\n{'='*80}")

    try:
        async with async_session_maker() as session:
            company_id_list = await _company_ids_for_batch(session, import_batch_id)
            print(f"[UPLOAD] Batch resolved to {len(company_id_list)} companies "
                  f"(via permanent company_import_batch membership table)")

            if await _cancel_requested(session, import_batch_id):
                await session.execute(
                    update(IcpImportBatch)
                    .where(IcpImportBatch.import_batch_id == import_batch_id)
                    .values(
                        scoring_status="complete",
                        research_status="complete_with_warnings",
                        processing_started_at=started,
                        processing_completed_at=datetime.now(timezone.utc),
                        processing_warnings=["Cancelled before processing started."],
                    )
                )
                await session.commit()
                return

            # Ensure a real Offering Profile (auto-sync if none/stale - item 11).
            print(f"[UPLOAD] Ensuring XSparks Offering Profile is fresh (auto-syncs from "
                  f"xsparks.ai if missing/stale)...")
            await ensure_offering_profile(session, organisation_id)
            # Domain enrichment is deliberately NOT run here. company_enrichment
            # still exists and is tested, but it made every upload block on one
            # web search per domain-less company (2,552 on a single file) with
            # no DB write until it finished, so the whole job looked frozen with
            # every company stuck at 'queued'. Re-enable by calling
            # enrich_missing_domains here - ideally after research rather than
            # before it, so the visible stages start immediately.
            #
            # Consequence while it is off: research skips any company whose
            # company_domain is NULL, so a spreadsheet with no website column
            # scores on contact access alone. That is surfaced as a warning at
            # ingest time (table_mapper.report_warnings) rather than hidden.
            domainless = sum(
                1 for row in (
                    await session.execute(
                        select(Company.company_id).where(
                            Company.company_id.in_(company_id_list),
                            Company.company_domain.is_(None),
                        )
                    )
                ).all()
            )
            if domainless:
                warnings.append(
                    f"{domainless} of {len(company_id_list)} company(ies) have no website - "
                    "research still runs on company name, but match confidence is lower without a domain"
                )
                print(f"[UPLOAD] {domainless}/{len(company_id_list)} companies have no domain - "
                      f"research will use company name (domain enrichment is disabled)")

            print(f"[UPLOAD] Offering Profile ready. Handing off to research_companies() "
                  f"for {len(company_id_list)} companies...")

            research_start = datetime.now(timezone.utc)
            research_summary = await search_signal_ingest.research_companies(
                session, organisation_id, company_ids=company_id_list, import_batch_id=import_batch_id,
            )
            research_elapsed = (datetime.now(timezone.utc) - research_start).total_seconds()
            print(f"\n[UPLOAD] === RESEARCH STAGE COMPLETE in {research_elapsed:.1f}s ===")
            print(f"[UPLOAD]   researched:        {research_summary.get('researched', 0)}")
            print(f"[UPLOAD]   successful:        {research_summary.get('successful', 0)}")
            print(f"[UPLOAD]   failed:            {research_summary.get('failed', 0)}")
            print(f"[UPLOAD]   research_failures: {research_summary.get('research_failures', 0)} (Tavily)")
            print(f"[UPLOAD]   llm_failures:      {research_summary.get('llm_failures', 0)}")
            print(f"[UPLOAD]   events_stored:     {research_summary.get('events_stored', 0)}\n")

            # Firmographics (industry / HQ / revenue / funding) from you.com.
            # Runs AFTER buying-event research so the visible research stage
            # starts immediately, and BEFORE scoring so Expected Deal Value can
            # use a newly found revenue figure. Only NULL columns are written.
            firmographic_start = datetime.now(timezone.utc)
            firmographic_summary = await company_enrichment.enrich_missing_firmographics(
                session, organisation_id, company_ids=company_id_list,
            )
            firmographic_elapsed = (datetime.now(timezone.utc) - firmographic_start).total_seconds()
            print(f"[UPLOAD] === FIRMOGRAPHICS STAGE COMPLETE in {firmographic_elapsed:.1f}s ===")
            print(f"[UPLOAD]   attempted: {firmographic_summary.get('attempted', 0)}")
            print(f"[UPLOAD]   updated:   {firmographic_summary.get('updated', 0)}")
            print(f"[UPLOAD]   failed:    {firmographic_summary.get('failed', 0)}\n")
            if firmographic_summary.get("failed"):
                warnings.append(
                    f"{firmographic_summary['failed']} company(ies) failed firmographic enrichment"
                )

            if await _cancel_requested(session, import_batch_id):
                await session.execute(
                    update(IcpImportBatch)
                    .where(IcpImportBatch.import_batch_id == import_batch_id)
                    .values(
                        scoring_status="complete",
                        research_status="complete_with_warnings",
                        companies_researched=research_summary.get("successful", 0),
                        research_failure_count=research_summary.get("research_failures", 0),
                        llm_failure_count=research_summary.get("llm_failures", 0),
                        processing_started_at=started,
                        processing_completed_at=datetime.now(timezone.utc),
                        processing_warnings=["Cancelled after research, before scoring."],
                    )
                )
                await session.commit()
                return

            print(f"[UPLOAD] Handing off to run_scoring() for {len(company_id_list)} companies...")
            scoring_start = datetime.now(timezone.utc)
            counts = await evidence_scorer.run_scoring(
                session, organisation_id, company_ids=company_id_list, import_batch_id=import_batch_id,
            )
            scoring_elapsed = (datetime.now(timezone.utc) - scoring_start).total_seconds()
            print(f"\n[UPLOAD] === SCORING STAGE COMPLETE in {scoring_elapsed:.1f}s ===")
            print(f"[UPLOAD]   Sales Ready:   {counts.get('Sales Ready', 0)}")
            print(f"[UPLOAD]   High Priority: {counts.get('High Priority', 0)}")
            print(f"[UPLOAD]   Warm:          {counts.get('Warm', 0)}")
            print(f"[UPLOAD]   Monitor:       {counts.get('Monitor', 0)}")
            print(f"[UPLOAD]   Low Priority:  {counts.get('Low Priority', 0)}\n")

            signals_extracted = (
                await session.execute(
                    select(func.count())
                    .select_from(BuyingEvent)
                    .join(CompanyImportBatch, CompanyImportBatch.company_id == BuyingEvent.company_id)
                    .where(CompanyImportBatch.import_batch_id == import_batch_id)
                )
            ).scalar() or 0

            failures = research_summary.get("failed", 0)
            if failures:
                warnings.append(f"{failures} company(ies) failed research (web search/LLM unavailable)")
            if research_summary.get("search_not_configured"):
                warnings.append("Web search not configured - no external evidence gathered")
            status = "complete_with_warnings" if warnings else "complete"

            await session.execute(
                update(IcpImportBatch)
                .where(IcpImportBatch.import_batch_id == import_batch_id)
                .values(
                    signals_extracted=signals_extracted,
                    companies_researched=research_summary.get("successful", 0),
                    research_failure_count=research_summary.get("research_failures", 0),
                    llm_failure_count=research_summary.get("llm_failures", 0),
                    scoring_failure_count=counts.get("_failures", 0),
                    sales_ready_count=counts["Sales Ready"],
                    high_priority_count=counts["High Priority"],
                    warm_count=counts["Warm"],
                    monitor_count=counts["Monitor"],
                    low_priority_count=counts["Low Priority"],
                    scoring_status="complete",
                    research_status=status,
                    processing_started_at=started,
                    processing_completed_at=datetime.now(timezone.utc),
                    processing_warnings=warnings or None,
                )
            )
            await session.commit()
        logger.info("job finished", extra={**log_ctx, "status": status})
    except Exception as exc:  # never leave the batch permanently pending
        error = f"{type(exc).__name__}: {exc}"[:1000]
        logger.exception("job failed", extra=log_ctx)
        try:
            async with async_session_maker() as session:
                await session.execute(
                    update(IcpImportBatch)
                    .where(IcpImportBatch.import_batch_id == import_batch_id)
                    .values(
                        scoring_status="complete",
                        research_status="failed",
                        processing_started_at=started,
                        processing_completed_at=datetime.now(timezone.utc),
                        processing_error=error,
                        processing_warnings=warnings or None,
                    )
                )
                await session.commit()
        except Exception:
            pass


async def _refresh_batch_sales_status_counts(session: AsyncSession, import_batch_id: UUID) -> None:
    """Re-tallies the WHOLE batch's sales-status counts from current
    LeadScore state - used after a partial retry, since the retry only
    reprocesses the failed subset but the batch's aggregate counts should
    reflect all of its companies, not just the ones just retried."""
    company_ids_subq = select(CompanyImportBatch.company_id).where(
        CompanyImportBatch.import_batch_id == import_batch_id
    )
    rows = (
        await session.execute(select(LeadScore.sales_status).where(LeadScore.company_id.in_(company_ids_subq)))
    ).scalars().all()
    counts = {label: 0 for label in evidence_scorer.STATUS_TO_COUNT}
    for status in rows:
        if status in counts:
            counts[status] += 1
    await session.execute(
        update(IcpImportBatch)
        .where(IcpImportBatch.import_batch_id == import_batch_id)
        .values(
            sales_ready_count=counts["Sales Ready"],
            high_priority_count=counts["High Priority"],
            warm_count=counts["Warm"],
            monitor_count=counts["Monitor"],
            low_priority_count=counts["Low Priority"],
        )
    )


async def retry_failed_companies_in_background(
    organisation_id: UUID, import_batch_id: UUID, company_ids: list[UUID]
) -> None:
    """POST .../retry-failed's background work: re-runs research+scoring for
    exactly the given (previously-failed, non-permanent) companies, then
    refreshes the whole batch's aggregate sales-status counts. Companies not
    in `company_ids` are untouched - a retry never re-processes already-
    completed work (brief: idempotent, no duplicate processing)."""
    log_ctx = {"job_id": str(import_batch_id), "stage": "retry"}
    logger.info("retry started, %d companies", len(company_ids), extra=log_ctx)
    try:
        async with async_session_maker() as session:
            await ensure_offering_profile(session, organisation_id)
            await search_signal_ingest.research_companies(
                session, organisation_id, company_ids=company_ids, import_batch_id=import_batch_id,
            )
            await evidence_scorer.run_scoring(
                session, organisation_id, company_ids=company_ids, import_batch_id=import_batch_id,
            )
            await _refresh_batch_sales_status_counts(session, import_batch_id)
            await session.commit()
        logger.info("retry finished", extra=log_ctx)
    except Exception:
        logger.exception("retry failed", extra=log_ctx)


async def record_import_batch(
    session: AsyncSession,
    workspace_id: UUID,
    file_names: list[str],
    total_rows: int,
    zi_to_company_id: dict[int, UUID],
    ingest_warnings: list[str] | None = None,
) -> IcpImportBatch:
    """Persists one prospect-upload event for the Settings prospect-data
    history (and Enterprise List's per-upload filter) - a permanent audit
    record, workspace-scoped, with NO ICP (brief section 7).

    Created with research_status='pending' and zero counts - research +
    scoring run afterward as a background task (score_companies_in_background),
    which fills the counts and flips the status. Records membership in
    company_import_batch (the permanent M:N table - brief item 5) so a later
    re-upload never erases this batch's membership, and also stamps the legacy
    Company.import_batch_id for the "most recent upload" convenience view."""
    batch = IcpImportBatch(
        workspace_id=workspace_id,
        icp_id=None,
        file_names=file_names,
        files_processed=len(file_names),
        total_rows=total_rows,
        companies_ingested=len(zi_to_company_id),
        signals_extracted=0,
        matched_icp_count=0,  # legacy column, unused by the new pipeline
        active_count=0,
        nurture_count=0,
        scoring_status="pending",
        research_status="pending",
        # Recorded at creation, not at completion: a file whose rows were all
        # unreadable finishes in milliseconds with nothing to research, so the
        # warning has to be attached here or the batch reads "Complete" with no
        # hint that 1,009 rows were discarded.
        processing_warnings=ingest_warnings or None,
    )
    session.add(batch)
    await session.commit()
    await session.refresh(batch)

    company_ids = list(zi_to_company_id.values())
    if company_ids:
        # Permanent membership (item 5) - one row per (company, batch), never
        # overwritten by a later upload.
        await session.execute(
            pg_insert(CompanyImportBatch)
            .values([{"company_id": cid, "import_batch_id": batch.import_batch_id} for cid in company_ids])
            .on_conflict_do_nothing(index_elements=["company_id", "import_batch_id"])
        )
        # Legacy "most recent upload" pointer, kept for convenience only.
        await session.execute(
            update(Company)
            .where(Company.company_id.in_(company_ids))
            .values(import_batch_id=batch.import_batch_id)
        )
        await session.commit()

    return batch


async def list_import_batches(session: AsyncSession, workspace_id: UUID) -> list[IcpImportBatch]:
    """Every prospect upload in this workspace, newest first. Workspace-scoped
    directly now (icp_import_batch.workspace_id) - no ICP join."""
    stmt = (
        select(IcpImportBatch)
        .where(IcpImportBatch.workspace_id == workspace_id)
        .order_by(IcpImportBatch.created_at.desc())
    )
    return (await session.execute(stmt)).scalars().all()


# Evidence-based export schema (brief section 23). No ICP/gate/D1-D7/component
# columns - those are gone from the active product.
EXPORT_COLUMNS = [
    "Company Name", "Domain", "Industry", "Location", "Employees", "Revenue",
    "Lead Score", "Sales Status", "Confidence", "Buying Evidence", "Contact Access",
    "Negative Penalty", "Best XSparks Offering", "Why Now", "Recommended Action",
    "Expected Deal Min", "Expected Deal Max", "Expected Deal Value", "Deal Value Basis",
    "Last Scored",
    # Added so a row explains ITSELF. Previously a company with no domain
    # exported as a name and a score of 20 with no indication that it scored low
    # because it was never researched, rather than because it was researched and
    # found uninteresting - two very different things for a rep working the sheet.
    "Researched", "Why Not Researched", "Buying Events", "Top Event", "Evidence Summary",
    "Contacts", "Primary Contact", "Primary Contact Title", "Primary Contact Email",
    "Scoring Warnings",
]

CONTACT_COLUMNS = [
    "Company Name", "First Name", "Last Name", "Job Title", "Department", "Persona",
    "Email", "Phone", "Mobile", "LinkedIn",
]

EVENT_COLUMNS = [
    "Company Name", "Event Type", "Category", "Title", "Summary", "Published",
    "Event Score", "Base Strength", "Relevance", "Freshness", "Source Quality",
    "Confidence", "Negative", "Penalty", "Best Offering", "Source URLs",
]


def _num(value) -> float | None:
    return float(value) if value is not None else None


def _contact_rank(contact) -> tuple:
    """Sort key picking the contact a rep would actually call: a verified email
    first, then seniority. Mirrors CONTACT_ACCESS in scoring_config, which
    scores a company on its single strongest reachable contact."""
    seniority = ("ceo", "coo", "cto", "cio", "chief", "founder", "president", "vp", "director")
    title = (contact.job_title or "").lower()
    rank = next((i for i, word in enumerate(seniority) if word in title), len(seniority))
    return (0 if contact.email else 1, rank, contact.last_name or "")


def _autosize(ws, max_width: int = 60) -> None:
    """Column widths from the content. Without this every column is 8 characters
    wide and the sheet is unreadable until the recipient resizes 30 of them."""
    for column in ws.columns:
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max(12, longest + 2), max_width)


def _format_evidence_summary(value) -> str | None:
    """LeadScore.evidence_summary is JSONB - a LIST of event dicts, not prose.
    openpyxl raises "Cannot convert [...] to Excel" on a list, so it has to be
    rendered here. One event per line, strongest first, in the same shape a rep
    reads on the Score Breakdown page."""
    if not value:
        return None
    if isinstance(value, str):
        return value
    if not isinstance(value, list):
        return str(value)
    lines = []
    for item in value:
        if not isinstance(item, dict):
            lines.append(str(item))
            continue
        score = item.get("event_score")
        parts = [f"[{float(score):.1f}]" if score is not None else "[-]"]
        if item.get("event_type"):
            parts.append(str(item["event_type"]))
        text = item.get("title") or item.get("summary")
        if text:
            parts.append(str(text))
        sources = item.get("sources")
        if sources:
            parts.append(f"({sources} source{'s' if sources != 1 else ''})")
        lines.append(" ".join(parts))
    return chr(10).join(lines) or None


def _text_list(value) -> str | None:
    """TEXT[] / list column rendered for a cell; tolerates a bare string."""
    if not value:
        return None
    if isinstance(value, str):
        return value
    return "; ".join(str(v) for v in value)


def build_company_export_workbook(
    rows: list[tuple[Company, LeadScore | None]],
    contacts: list | None = None,
    events: list | None = None,
) -> bytes:
    """Enterprise List "Export" - three sheets, so the workbook carries the
    whole picture rather than one score per company:

      Companies     one row each, with the score AND why it is what it is
      Contacts      every decision maker, with company
      Buying Events every live event, with its full score breakdown

    contacts/events are optional so existing callers keep working; when absent
    only the Companies sheet is written."""
    contacts = contacts or []
    events = events or []

    by_company_contacts: dict = {}
    for contact, company_name in contacts:
        by_company_contacts.setdefault(contact.company_id, []).append((contact, company_name))
    by_company_events: dict = {}
    for event, company_name in events:
        if not event.is_negative:
            by_company_events.setdefault(event.company_id, []).append(event)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"
    ws.append(EXPORT_COLUMNS)

    for company, score in rows:
        location = ", ".join(p for p in [company.city, company.state, company.country] if p) or None
        base = [
            company.company_name, company.company_domain,
            (company.industries or [None])[0] if company.industries else None,
            location, company.employee_count, company.revenue_usd,
        ]
        if score is None:
            score_values = [None] * 14
        else:
            score_values = [
                _num(score.lead_score), score.sales_status, score.confidence_label,
                _num(score.buying_evidence_score), _num(score.contact_access_score),
                _num(score.negative_event_score), score.best_offering, score.why_now,
                score.recommended_action, _num(score.expected_deal_min_usd),
                _num(score.expected_deal_max_usd), _num(score.expected_deal_value_usd),
                score.deal_value_basis,
                score.scored_at.isoformat() if score.scored_at else None,
            ]

        company_events = sorted(
            by_company_events.get(company.company_id, []),
            key=lambda e: float(e.event_score or 0), reverse=True,
        )
        researched = company.search_signals_fetched_at is not None
        # The distinction that makes a low score readable: never researched vs
        # researched and genuinely quiet.
        if researched:
            not_researched_reason = None
        elif not company.company_domain:
            not_researched_reason = "No website - research needs a domain"
        else:
            not_researched_reason = "Not yet researched"

        ranked_contacts = sorted(
            (c for c, _ in by_company_contacts.get(company.company_id, [])), key=_contact_rank
        )
        primary = ranked_contacts[0] if ranked_contacts else None

        context_values = [
            "Yes" if researched else "No",
            not_researched_reason,
            len(company_events),
            company_events[0].title if company_events else None,
            _format_evidence_summary(score.evidence_summary) if score is not None else None,
            len(ranked_contacts),
            " ".join(p for p in [primary.first_name, primary.last_name] if p) if primary else None,
            primary.job_title if primary else None,
            primary.email if primary else None,
            _text_list(score.scoring_warnings) if score is not None else None,
        ]
        ws.append(base + score_values + context_values)
    _autosize(ws)

    if contacts:
        cs = wb.create_sheet("Contacts")
        cs.append(CONTACT_COLUMNS)
        for contact, company_name in contacts:
            cs.append([
                company_name, contact.first_name, contact.last_name, contact.job_title,
                contact.department, contact.persona, contact.email, contact.phone,
                contact.mobile_phone, contact.linkedin_url,
            ])
        _autosize(cs)

    if events:
        es = wb.create_sheet("Buying Events")
        es.append(EVENT_COLUMNS)
        for event, company_name in events:
            urls = [e.get("url") for e in (event.evidence or []) if isinstance(e, dict) and e.get("url")]
            es.append([
                company_name, event.event_type, event.category, event.title, event.summary,
                event.published_at.isoformat() if event.published_at else None,
                _num(event.event_score), _num(event.base_strength), _num(event.relevance),
                _num(event.freshness), _num(event.source_quality), _num(event.extraction_confidence),
                "Yes" if event.is_negative else "No", _num(event.penalty_value),
                event.best_offering, "\n".join(urls) or None,
            ])
        _autosize(es)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def delete_import_batch(
    session: AsyncSession, workspace_id: UUID, import_batch_id: UUID
) -> dict | None:
    """Deletes one upload and the data it introduced. Returns counts, or None
    when the batch does not exist in this workspace.

    The correctness risk here is SHARED COMPANIES. company_import_batch is a
    permanent many-to-many (brief item 5): a company that appeared in three
    uploads has three membership rows, and 64 such rows exist today. Deleting
    a batch must therefore remove only the companies whose ONLY membership is
    this batch - anything also present in another upload keeps its row there
    and must survive, or deleting an old upload would silently destroy
    companies a newer one still relies on.

    Everything hanging off a deleted company (buying_event, lead_score,
    decision_maker, trigger_event, company_import_batch, and the legacy
    intent/news/scoop/signal rows) is removed by ON DELETE CASCADE - see
    migration d1a7f3c8e5b2, which closed the last two FKs that were still
    NO ACTION and blocking exactly this.
    """
    batch = (
        await session.execute(
            select(IcpImportBatch).where(
                IcpImportBatch.import_batch_id == import_batch_id,
                IcpImportBatch.workspace_id == workspace_id,
            )
        )
    ).scalar_one_or_none()
    if batch is None:
        return None  # absent, or another workspace's - caller turns this into a 404

    exclusive_ids = [
        row[0]
        for row in (
            await session.execute(
                text(
                    """
                    SELECT cib.company_id
                    FROM company_import_batch cib
                    WHERE cib.import_batch_id = :batch_id
                      AND NOT EXISTS (
                          SELECT 1 FROM company_import_batch other
                          WHERE other.company_id = cib.company_id
                            AND other.import_batch_id <> :batch_id
                      )
                    """
                ),
                {"batch_id": import_batch_id},
            )
        ).all()
    ]

    total_members = (
        await session.execute(
            select(func.count())
            .select_from(CompanyImportBatch)
            .where(CompanyImportBatch.import_batch_id == import_batch_id)
        )
    ).scalar_one()

    events_removed = 0
    if exclusive_ids:
        events_removed = (
            await session.execute(
                select(func.count())
                .select_from(BuyingEvent)
                .where(BuyingEvent.company_id.in_(exclusive_ids))
            )
        ).scalar_one()
        # Cascades do the rest. Chunked because a large upload can exceed the
        # parameter limit of a single IN (...) statement.
        for start in range(0, len(exclusive_ids), 500):
            chunk = exclusive_ids[start : start + 500]
            await session.execute(delete(Company).where(Company.company_id.in_(chunk)))

    # Deleting the batch cascades its remaining company_import_batch rows,
    # which is what detaches the shared companies without harming them.
    await session.delete(batch)
    await session.commit()

    return {
        "import_batch_id": str(import_batch_id),
        "file_names": batch.file_names or [],
        "companies_deleted": len(exclusive_ids),
        "companies_kept": total_members - len(exclusive_ids),
        "buying_events_deleted": events_removed,
    }
